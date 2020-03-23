import openpyxl
import smtplib
import ssl
import getpass

loc = "NCC.xlsx"

smtp_server = "smtp.gmail.com"
port = 465
sender = "credenzuser@gmail.com"
passwd = getpass.getpass("Enter your password:- ")

wb = openpyxl.load_workbook(loc)
sheet = wb.active
rows = sheet.max_row
context = ssl.create_default_context()


context = ssl.create_default_context()
with smtplib.SMTP_SSL(smtp_server, port) as server:
    server.login(sender, passwd)
    for i in range(2, rows+1):
        r_email = str(sheet.cell(row=i, column=4).value)
        name1 = str(sheet.cell(row=i, column=2).value)
        name2 = str(sheet.cell(row=i, column=3).value)        
        username = sheet.cell(row=i, column=7).value
        password = sheet.cell(row=i, column=8).value
        subject = "Here are your login credentials"
        if name2 == 'None' or name2=='':
        	body = f"Hi {name1}!\nThank you for registering in the National Computing Contest 2020.\n\nHere are your login credentials:\nUsername - {username}\nPassword - {password}\n\nBefore logging in, it would be beneficial to go over some general instructions:\n1. The contest is best viewed in full-screen and 1366x768 resolution settings. We suggest that you change your display settings temporarily, in order to enhance your experience.\n2. Make sure you have a stable internet connection for the next two hours.\n3. Refreshing the page at any point during the game will cause loss of your data, which cannot be retrieved. We thereby advice you to abstain from such practices.\n\nPlease share these credentials with your teammate as well. Happy coding!\nSee you on the leaderboard!"
        	message = f'Subject: {subject}\n\n{body}' 
        else:
        	body = f"Hi {name1} and {name2}!\nThank you for registering in the National Computing Contest 2020.\n\nHere are your login credentials:\nUsername - {username}\nPassword - {password}\n\nBefore logging in, it would be beneficial to go over some general instructions:\n1. The contest is best viewed in full-screen and 1366x768 resolution settings. We suggest that you change your display settings temporarily, in order to enhance your experience.\n2. Make sure you have a stable internet connection for the next two hours.\n3. Refreshing the page at any point during the game will cause loss of your data, which cannot be retrieved. We thereby advice you to abstain from such practices.\n\nPlease share these credentials with your teammate as well. Happy coding!\nSee you on the leaderboard!"
        	message = f'Subject: {subject}\n\n{body}'
        
        server.sendmail(
            sender, r_email, message
        )
        print("Email sent to "+r_email)
        
        
        
        
        
        
        
        
