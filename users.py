import openpyxl
import datetime
import time
import random


loc = "NCC.xlsx"
words = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789!@#$%^&*"


wb = openpyxl.load_workbook(loc)
sheet = wb.active
rows = sheet.max_row

for i in range(2, rows+1):
    x = datetime.datetime.now()
    name = sheet.cell(row=i, column=2).value
    category = sheet.cell(row=i, column=5).value
    username = category[0]+x.strftime("%M%S%f")[:-2]
    time.sleep(random.random())
    temp = sheet.cell(row=i, column=7)
    temp.value = username
    password = ""
    for j in range(0, 8):
        password = password + random.choice(words)
    temp = sheet.cell(row=i, column=8)
    temp.value = password
    print(f'Iteration number:- {i-1}')


wb.save(loc)
